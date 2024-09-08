import openpyxl
import os
from task_model import Task, Subtask  # Import Task and Subtask from task_model.py

class ExcelHandler:
    def __init__(self, excel_file: str = "task_manager_data.xlsx"):
        """
        Initialize the ExcelHandler with the specified Excel file.
        If no file is provided, it defaults to 'task_manager_data.xlsx'.
        """
        self.excel_file = excel_file

    def load_data(self) -> tuple[list[Task], list[Subtask]]:
        """
        Load tasks and subtasks from the Excel file.
        If the file does not exist, it creates a new Excel file with the required structure.
        :return: A tuple containing a list of Task objects and a list of Subtask objects.
        """
        if not os.path.exists(self.excel_file):
            self.create_excel_file()  # Create Excel file if it doesn't exist

        workbook = openpyxl.load_workbook(self.excel_file)

        # Load tasks from the 'Main Tasks' sheet
        main_task_sheet = workbook["Main Tasks"]
        tasks = [
            Task(*[cell.value for cell in row]) for row in main_task_sheet.iter_rows(min_row=2)
        ]

        # Load subtasks from the 'Subtasks' sheet
        subtask_sheet = workbook["Subtasks"]
        subtasks = [
            Subtask(*[cell.value for cell in row]) for row in subtask_sheet.iter_rows(min_row=2)
        ]

        return tasks, subtasks

    def create_excel_file(self):
        """
        Create a new Excel file with the necessary sheets and headers.
        This method sets up two sheets: 'Main Tasks' for tasks and 'Subtasks' for subtasks.
        """
        workbook = openpyxl.Workbook()

        # Create the 'Main Tasks' sheet and add headers
        main_task_sheet = workbook.active
        main_task_sheet.title = "Main Tasks"
        main_task_headers = ["Task ID", "Task Name", "Category", "Priority", "Start Date", "Due Date", "Status", "Progress", "Notes"]
        main_task_sheet.append(main_task_headers)

        # Create the 'Subtasks' sheet and add headers
        subtask_sheet = workbook.create_sheet(title="Subtasks")
        subtask_headers = ["Subtask ID", "Task ID", "Subtask Name", "Subtask Status", "Subtask Progress", "Subtask Due Date", "Subtask Completed Date"]
        subtask_sheet.append(subtask_headers)

        # Save the new Excel file
        workbook.save(self.excel_file)

    def save_data(self, tasks: list[Task], subtasks: list[Subtask]):
        """
        Save tasks and subtasks to the Excel file.
        The existing 'Main Tasks' and 'Subtasks' sheets are cleared and repopulated with new data.
        :param tasks: List of Task objects to be saved.
        :param subtasks: List of Subtask objects to be saved.
        """
        workbook = openpyxl.load_workbook(self.excel_file)

        # Remove the old 'Main Tasks' sheet if it exists and create a new one
        if "Main Tasks" in workbook.sheetnames:
            main_task_sheet = workbook["Main Tasks"]
            workbook.remove(main_task_sheet)
        main_task_sheet = workbook.create_sheet(title="Main Tasks")

        # Add headers to the new 'Main Tasks' sheet
        main_task_headers = ["Task ID", "Task Name", "Category", "Priority", "Start Date", "Due Date", "Status", "Progress", "Notes"]
        main_task_sheet.append(main_task_headers)

        # Add task data to the 'Main Tasks' sheet
        for task in tasks:
            main_task_sheet.append(task.as_list())

        # Remove the old 'Subtasks' sheet if it exists and create a new one
        if "Subtasks" in workbook.sheetnames:
            subtask_sheet = workbook["Subtasks"]
            workbook.remove(subtask_sheet)
        subtask_sheet = workbook.create_sheet(title="Subtasks")

        # Add headers to the new 'Subtasks' sheet
        subtask_headers = ["Subtask ID", "Task ID", "Subtask Name", "Subtask Status", "Subtask Progress", "Subtask Due Date", "Subtask Completed Date"]
        subtask_sheet.append(subtask_headers)

        # Add subtask data to the 'Subtasks' sheet
        for subtask in subtasks:
            subtask_sheet.append(subtask.as_list())

        # Save the workbook with the updated tasks and subtasks
        workbook.save(self.excel_file)
