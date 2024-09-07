import openpyxl
from tkinter import messagebox
import os


class TaskManagerLogic:
    def __init__(self):
        self.excel_file = "task_manager_data.xlsx"
        self.main_tasks = []
        self.subtasks = []

        # Load tasks and subtasks from the Excel file on startup
        self.load_from_excel()

    def load_from_excel(self):
        """Loads tasks and subtasks from an Excel file."""
        if not os.path.exists(self.excel_file):
            # Create the Excel file if it doesn't exist
            self.create_excel_file()

        workbook = openpyxl.load_workbook(self.excel_file)

        # Load main tasks from the "Main Tasks" sheet
        main_task_sheet = workbook["Main Tasks"]
        self.main_tasks = [
            [cell.value for cell in row] for row in main_task_sheet.iter_rows(min_row=2)
        ]

        # Load subtasks from the "Subtasks" sheet
        subtask_sheet = workbook["Subtasks"]
        self.subtasks = [
            [cell.value for cell in row] for row in subtask_sheet.iter_rows(min_row=2)
        ]

    def create_excel_file(self):
        """Creates a new Excel file with the necessary sheets and headers."""
        workbook = openpyxl.Workbook()

        # Create the main task sheet with headers
        main_task_sheet = workbook.active
        main_task_sheet.title = "Main Tasks"
        main_task_headers = ["Task ID", "Task Name", "Category", "Priority", "Start Date", "Due Date", "Status", "Progress", "Notes"]
        main_task_sheet.append(main_task_headers)

        # Create the subtask sheet with headers
        subtask_sheet = workbook.create_sheet(title="Subtasks")
        subtask_headers = ["Subtask ID", "Task ID", "Subtask Name", "Subtask Status", "Subtask Progress", "Subtask Due Date", "Subtask Completed Date"]
        subtask_sheet.append(subtask_headers)

        # Save the workbook
        workbook.save(self.excel_file)

    def save_to_excel(self):
        """Saves tasks and subtasks to an Excel file."""
        workbook = openpyxl.load_workbook(self.excel_file)

        # Clear and re-populate the main task sheet
        if "Main Tasks" in workbook.sheetnames:
            main_task_sheet = workbook["Main Tasks"]
            workbook.remove(main_task_sheet)
        main_task_sheet = workbook.create_sheet(title="Main Tasks")

        # Add headers to the main task sheet
        main_task_headers = ["Task ID", "Task Name", "Category", "Priority", "Start Date", "Due Date", "Status", "Progress", "Notes"]
        main_task_sheet.append(main_task_headers)

        # Add task data to the main task sheet
        for task in self.main_tasks:
            main_task_sheet.append(task)

        # Clear and re-populate the subtask sheet
        if "Subtasks" in workbook.sheetnames:
            subtask_sheet = workbook["Subtasks"]
            workbook.remove(subtask_sheet)
        subtask_sheet = workbook.create_sheet(title="Subtasks")

        # Add headers to the subtask sheet
        subtask_headers = ["Subtask ID", "Task ID", "Subtask Name", "Subtask Status", "Subtask Progress", "Subtask Due Date", "Subtask Completed Date"]
        subtask_sheet.append(subtask_headers)

        # Add subtask data to the subtask sheet
        for subtask in self.subtasks:
            subtask_sheet.append(subtask)

        # Save the workbook
        workbook.save(self.excel_file)

    def get_new_task_id(self):
        """Generates a new unique Task ID."""
        if not self.main_tasks:
            return 1
        return max(task[0] for task in self.main_tasks) + 1

    def get_new_subtask_id(self):
        """Generates a new unique Subtask ID."""
        if not self.subtasks:
            return 1
        return max(subtask[0] for subtask in self.subtasks) + 1

    def add_task(self, task_data):
        """Adds a task to the main tasks."""
        task_id = self.get_new_task_id()  # Generate a unique Task ID
        new_task = [
            task_id,
            task_data['task_name'],
            task_data['category'],
            task_data['priority'],
            task_data['start_date'],
            task_data['due_date'],
            task_data['status'],
            task_data['progress'],
            task_data['notes']
        ]
        self.main_tasks.append(new_task)
        self.save_to_excel()  # Save changes to Excel

    def delete_task(self, task_id):
        """Deletes a task by ID."""
        self.main_tasks = [task for task in self.main_tasks if task[0] != task_id]
        self.subtasks = [subtask for subtask in self.subtasks if subtask[1] != task_id]  # Delete related subtasks
        self.save_to_excel()  # Save changes to Excel

    def delete_subtask(self, subtask_id):
        """Deletes a subtask by ID."""
        self.subtasks = [subtask for subtask in self.subtasks if subtask[0] != subtask_id]
        self.save_to_excel()  # Save changes to Excel

    def add_subtask(self, subtask_data):
        """Adds a subtask to the subtask list."""
        subtask_id = self.get_new_subtask_id()  # Generate a unique Subtask ID
        new_subtask = [
            subtask_id,
            subtask_data['task_id'],
            subtask_data['subtask_name'],
            subtask_data['subtask_status'],
            subtask_data['subtask_progress'],
            subtask_data['subtask_due_date'],
            subtask_data['subtask_completed_date']
        ]
        self.subtasks.append(new_subtask)
        self.save_to_excel()  # Save changes to Excel
